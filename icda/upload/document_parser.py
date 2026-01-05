"""
ICDA Document Parser Service
============================
Parses various file formats into normalized records for indexing/validation.
Handles streaming for large files.

Author: Bishop Walker / Salt Water Coder
Project: ICDA Prototype
"""

import asyncio
import csv
import io
import json
import re
from abc import ABC, abstractmethod
from dataclasses import dataclass
from pathlib import Path
from typing import Any, AsyncIterator, Iterator, Optional
import logging

logger = logging.getLogger(__name__)


@dataclass
class ParsedRecord:
    """A single parsed record from any file type"""
    row_number: int
    data: dict
    raw_text: Optional[str] = None
    metadata: dict = None
    
    def __post_init__(self):
        if self.metadata is None:
            self.metadata = {}


class BaseParser(ABC):
    """Abstract base for all file parsers"""
    
    @abstractmethod
    async def parse(self, file_path: Path) -> list["ParsedRecord"]:
        """Parse file and return list of records"""
        pass
    
    @abstractmethod
    async def stream_parse(self, file_path: Path) -> AsyncIterator["ParsedRecord"]:
        """Stream parse for large files"""
        pass


class CSVParser(BaseParser):
    """
    CSV Parser with streaming support for large files.
    Handles various delimiters and encodings.
    """
    
    CHUNK_SIZE = 8192  # Read 8KB at a time
    
    def __init__(self, delimiter: str = None, encoding: str = "utf-8"):
        self.delimiter = delimiter
        self.encoding = encoding
    
    async def parse(self, file_path: Path) -> list[ParsedRecord]:
        """Parse entire CSV file"""
        records = []
        async for record in self.stream_parse(file_path):
            records.append(record)
        return records
    
    async def stream_parse(self, file_path: Path) -> AsyncIterator[ParsedRecord]:
        """Stream parse CSV for memory efficiency"""
        import aiofiles
        
        async with aiofiles.open(file_path, "r", encoding=self.encoding) as f:
            content = await f.read()
        
        # Auto-detect delimiter if not specified
        if not self.delimiter:
            self.delimiter = self._detect_delimiter(content[:1024])
        
        reader = csv.DictReader(io.StringIO(content), delimiter=self.delimiter)
        
        for row_num, row in enumerate(reader, start=1):
            # Clean up keys and values
            cleaned = {
                self._clean_key(k): self._clean_value(v)
                for k, v in row.items()
                if k is not None
            }
            
            yield ParsedRecord(
                row_number=row_num,
                data=cleaned,
                raw_text=self.delimiter.join(str(v) for v in row.values())
            )
    
    def _detect_delimiter(self, sample: str) -> str:
        """Auto-detect CSV delimiter"""
        delimiters = [",", "\t", "|", ";"]
        counts = {d: sample.count(d) for d in delimiters}
        return max(counts, key=counts.get)
    
    def _clean_key(self, key: str) -> str:
        """Normalize column names"""
        if not key:
            return "unknown"
        return re.sub(r"\s+", "_", key.strip().lower())
    
    def _clean_value(self, value: Any) -> Any:
        """Clean and normalize values"""
        if value is None:
            return None
        if isinstance(value, str):
            value = value.strip()
            if value.lower() in ("null", "none", "n/a", "na", ""):
                return None
        return value


class ExcelParser(BaseParser):
    """
    Excel Parser supporting .xlsx and .xls formats.
    Uses openpyxl for xlsx, xlrd for xls.
    """
    
    def __init__(self, sheet_name: str = None):
        self.sheet_name = sheet_name
    
    async def parse(self, file_path: Path) -> list[ParsedRecord]:
        """Parse Excel file"""
        records = []
        async for record in self.stream_parse(file_path):
            records.append(record)
        return records
    
    async def stream_parse(self, file_path: Path) -> AsyncIterator[ParsedRecord]:
        """Stream parse Excel file"""
        import openpyxl
        
        # Run in thread pool for blocking I/O
        loop = asyncio.get_event_loop()
        wb = await loop.run_in_executor(
            None,
            lambda: openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        )
        
        try:
            # Select sheet
            if self.sheet_name and self.sheet_name in wb.sheetnames:
                ws = wb[self.sheet_name]
            else:
                ws = wb.active
            
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                return
            
            # First row is headers
            headers = [self._clean_key(str(h) if h else f"col_{i}") 
                      for i, h in enumerate(rows[0])]
            
            # Process data rows
            for row_num, row in enumerate(rows[1:], start=1):
                data = {}
                for i, (header, value) in enumerate(zip(headers, row)):
                    data[header] = self._clean_value(value)
                
                yield ParsedRecord(
                    row_number=row_num,
                    data=data,
                    raw_text=str(row)
                )
                
        finally:
            wb.close()
    
    def _clean_key(self, key: str) -> str:
        """Normalize column names"""
        if not key:
            return "unknown"
        return re.sub(r"\s+", "_", key.strip().lower())
    
    def _clean_value(self, value: Any) -> Any:
        """Clean and normalize values"""
        if value is None:
            return None
        if isinstance(value, str):
            value = value.strip()
            if value.lower() in ("null", "none", "n/a", "na", ""):
                return None
        return value


class JSONParser(BaseParser):
    """
    JSON Parser supporting objects, arrays, and JSONL.
    Handles nested structures by flattening.
    """
    
    def __init__(self, flatten: bool = True, max_depth: int = 3):
        self.flatten = flatten
        self.max_depth = max_depth
    
    async def parse(self, file_path: Path) -> list[ParsedRecord]:
        """Parse JSON file"""
        records = []
        async for record in self.stream_parse(file_path):
            records.append(record)
        return records
    
    async def stream_parse(self, file_path: Path) -> AsyncIterator[ParsedRecord]:
        """Stream parse JSON/JSONL file"""
        import aiofiles
        
        async with aiofiles.open(file_path, "r", encoding="utf-8") as f:
            content = await f.read()
        
        content = content.strip()
        
        # Detect if JSONL (newline-delimited JSON)
        if content.startswith("["):
            # Regular JSON array
            data = json.loads(content)
            if isinstance(data, list):
                for row_num, item in enumerate(data, start=1):
                    processed = self._flatten_dict(item) if self.flatten else item
                    yield ParsedRecord(
                        row_number=row_num,
                        data=processed,
                        raw_text=json.dumps(item)
                    )
            else:
                processed = self._flatten_dict(data) if self.flatten else data
                yield ParsedRecord(
                    row_number=1,
                    data=processed,
                    raw_text=json.dumps(data)
                )
        else:
            # Try JSONL
            row_num = 0
            for line in content.split("\n"):
                line = line.strip()
                if not line:
                    continue
                try:
                    item = json.loads(line)
                    row_num += 1
                    processed = self._flatten_dict(item) if self.flatten else item
                    yield ParsedRecord(
                        row_number=row_num,
                        data=processed,
                        raw_text=line
                    )
                except json.JSONDecodeError:
                    continue
    
    def _flatten_dict(self, d: dict, parent_key: str = "", sep: str = "_", depth: int = 0) -> dict:
        """Flatten nested dictionary"""
        if depth >= self.max_depth:
            return {parent_key: str(d)} if parent_key else {"value": str(d)}
        
        items = []
        for k, v in d.items():
            new_key = f"{parent_key}{sep}{k}" if parent_key else k
            if isinstance(v, dict):
                items.extend(self._flatten_dict(v, new_key, sep, depth + 1).items())
            elif isinstance(v, list):
                items.append((new_key, json.dumps(v) if v else None))
            else:
                items.append((new_key, v))
        return dict(items)


class PDFParser(BaseParser):
    """
    PDF Parser using pdfplumber for text extraction.
    Handles tables and text content.
    """
    
    def __init__(self, extract_tables: bool = True):
        self.extract_tables = extract_tables
    
    async def parse(self, file_path: Path) -> list[ParsedRecord]:
        """Parse PDF file"""
        records = []
        async for record in self.stream_parse(file_path):
            records.append(record)
        return records
    
    async def stream_parse(self, file_path: Path) -> AsyncIterator[ParsedRecord]:
        """Stream parse PDF file"""
        import pdfplumber
        
        loop = asyncio.get_event_loop()
        
        def extract_content():
            results = []
            with pdfplumber.open(file_path) as pdf:
                for page_num, page in enumerate(pdf.pages, start=1):
                    # Extract text
                    text = page.extract_text() or ""
                    
                    # Extract tables if enabled
                    tables = []
                    if self.extract_tables:
                        tables = page.extract_tables() or []
                    
                    results.append({
                        "page_num": page_num,
                        "text": text,
                        "tables": tables
                    })
            return results
        
        pages = await loop.run_in_executor(None, extract_content)
        
        row_num = 0
        for page_data in pages:
            # Yield text content as record
            if page_data["text"].strip():
                row_num += 1
                yield ParsedRecord(
                    row_number=row_num,
                    data={
                        "page": page_data["page_num"],
                        "content_type": "text",
                        "text": page_data["text"]
                    },
                    raw_text=page_data["text"],
                    metadata={"page": page_data["page_num"]}
                )
            
            # Yield table rows as separate records
            for table in page_data["tables"]:
                if not table or len(table) < 2:
                    continue
                
                headers = [str(h).strip().lower().replace(" ", "_") 
                          for h in table[0] if h]
                
                for table_row in table[1:]:
                    row_num += 1
                    data = {}
                    for i, val in enumerate(table_row):
                        if i < len(headers):
                            data[headers[i]] = val
                    
                    yield ParsedRecord(
                        row_number=row_num,
                        data=data,
                        raw_text=str(table_row),
                        metadata={"page": page_data["page_num"], "from_table": True}
                    )


class DOCXParser(BaseParser):
    """
    DOCX Parser using python-docx for text extraction.
    Handles paragraphs and tables.
    """
    
    def __init__(self, extract_tables: bool = True):
        self.extract_tables = extract_tables
    
    async def parse(self, file_path: Path) -> list[ParsedRecord]:
        """Parse DOCX file"""
        records = []
        async for record in self.stream_parse(file_path):
            records.append(record)
        return records
    
    async def stream_parse(self, file_path: Path) -> AsyncIterator[ParsedRecord]:
        """Stream parse DOCX file"""
        from docx import Document
        
        loop = asyncio.get_event_loop()
        doc = await loop.run_in_executor(None, Document, file_path)
        
        row_num = 0
        
        # Extract paragraphs
        for para in doc.paragraphs:
            if para.text.strip():
                row_num += 1
                yield ParsedRecord(
                    row_number=row_num,
                    data={
                        "content_type": "paragraph",
                        "text": para.text,
                        "style": para.style.name if para.style else None
                    },
                    raw_text=para.text
                )
        
        # Extract tables
        if self.extract_tables:
            for table_idx, table in enumerate(doc.tables):
                rows = []
                for row in table.rows:
                    rows.append([cell.text.strip() for cell in row.cells])
                
                if not rows:
                    continue
                
                headers = [h.lower().replace(" ", "_") for h in rows[0]]
                
                for table_row in rows[1:]:
                    row_num += 1
                    data = {}
                    for i, val in enumerate(table_row):
                        if i < len(headers):
                            data[headers[i]] = val
                    
                    yield ParsedRecord(
                        row_number=row_num,
                        data=data,
                        raw_text=str(table_row),
                        metadata={"table_index": table_idx, "from_table": True}
                    )


class TextParser(BaseParser):
    """
    Plain text parser for .txt, .md, etc.
    Splits into paragraphs or lines.
    """
    
    def __init__(self, split_mode: str = "paragraph"):
        """
        split_mode: 'paragraph' (double newline), 'line', or 'chunk' (by size)
        """
        self.split_mode = split_mode
        self.chunk_size = 1000  # characters per chunk
    
    async def parse(self, file_path: Path) -> list[ParsedRecord]:
        """Parse text file"""
        records = []
        async for record in self.stream_parse(file_path):
            records.append(record)
        return records
    
    async def stream_parse(self, file_path: Path) -> AsyncIterator[ParsedRecord]:
        """Stream parse text file"""
        import aiofiles
        
        async with aiofiles.open(file_path, "r", encoding="utf-8") as f:
            content = await f.read()
        
        if self.split_mode == "paragraph":
            chunks = re.split(r"\n\s*\n", content)
        elif self.split_mode == "line":
            chunks = content.split("\n")
        else:
            # Chunk by size
            chunks = [content[i:i+self.chunk_size] 
                     for i in range(0, len(content), self.chunk_size)]
        
        for row_num, chunk in enumerate(chunks, start=1):
            chunk = chunk.strip()
            if not chunk:
                continue
            
            yield ParsedRecord(
                row_number=row_num,
                data={
                    "content_type": "text",
                    "text": chunk
                },
                raw_text=chunk
            )


class DocumentParserService:
    """
    Main service for parsing documents.
    Routes to appropriate parser based on file type.
    """
    
    def __init__(self):
        self.parsers = {
            "csv": CSVParser(),
            "xlsx": ExcelParser(),
            "xls": ExcelParser(),
            "json": JSONParser(),
            "pdf": PDFParser(),
            "docx": DOCXParser(),
            "txt": TextParser(),
            "md": TextParser()
        }
    
    def get_parser(self, file_type: str) -> BaseParser:
        """Get appropriate parser for file type"""
        parser = self.parsers.get(file_type.lower())
        if not parser:
            raise ValueError(f"No parser available for file type: {file_type}")
        return parser
    
    async def parse_file(self, file_path: Path, file_type: str) -> list[dict]:
        """
        Parse a file and return list of record dictionaries.
        """
        parser = self.get_parser(file_type)
        records = await parser.parse(file_path)
        return [r.data for r in records]
    
    async def stream_parse_file(
        self,
        file_path: Path,
        file_type: str
    ) -> AsyncIterator[dict]:
        """
        Stream parse a file for memory efficiency.
        """
        parser = self.get_parser(file_type)
        async for record in parser.stream_parse(file_path):
            yield record.data
